import pandas as pd
import openpyxl


def get_datas():  # выгружает в df столбцы J : P
    wb = openpyxl.load_workbook('Приложение 1.xlsx', data_only=True)
    sheet = wb['Анализ_БК+ББ']
    customers_list = list([])
    usl_post = list([])
    data_otguzki = list([])
    dogovonoy_srok = list([])
    data_postup_viruchki = list([])
    for i in range(4, 29):
        customers_list.append(sheet[f"J{i}"].value)
        usl_post.append(sheet[f"K{i}"].value)
        data_otguzki.append(sheet[f"M{i}"].value)
        dogovonoy_srok.append(sheet[f"N{i}"].value)
        data_postup_viruchki.append(sheet[f"P{i}"].value)
    data = {'Покупатель': customers_list,
            'Условия поставок': usl_post,
            'Дата отгрузки': data_otguzki,
            'Договорной срок': dogovonoy_srok,
            'Дата поступления выручки': data_postup_viruchki}
    df = pd.DataFrame(data)
    df['Просрочка'] = df['Дата поступления выручки'] - df['Договорной срок']
    return df


def get_kurses():  # выгружает в df столбцы Y : AA
    wb = openpyxl.load_workbook('Приложение 1.xlsx', data_only=True)
    sheet = wb['Анализ_БК+ББ']
    kurs_r = list([])
    kurs_p = list([])
    realization = list([])
    for i in range(4, 29):
        kurs_r.append(sheet[f"Y{i}"].value)
        kurs_p.append(sheet[f"Z{i}"].value)
        realization.append(sheet[f"AA{i}"].value)
    data = {'Курс на дату реализации': kurs_r,
            'Курс на дату поступления денежных средств': kurs_p,
            'Реализация': realization}
    df = pd.DataFrame(data)
    return df


def get_AHAJ():
    wb = openpyxl.load_workbook('Приложение 1.xlsx', data_only=True)
    sheet = wb['Анализ_БК+ББ']
    rps = list([])
    r_usd = list([])
    r_roubles = list([])
    for i in range(4, 29):
        rps.append(sheet[f"AH{i}"].value)
        r_usd.append(sheet[f"AI{i}"].value)
        r_roubles.append(sheet[f"AJ{i}"].value)
    data = {'Реализация (предварительная сумма), USD': rps,
            'реализация, USD': r_usd,
            'реализация (отгрузка), рубли': r_roubles}
    df = pd.DataFrame(data)
    return df


def get_ARAS():
    wb = openpyxl.load_workbook('Приложение 1.xlsx', data_only=True)
    sheet = wb['Анализ_БК+ББ']
    pr = list([])
    krmo = list([])
    for i in range(4, 29):
        pr.append(sheet[f"AR{i}"].value)
        krmo.append(sheet[f"AS{i}"].value)
    data = {'Поступление, рубли': pr,
            'Курсовые разницы между отгрузкой и поступлением денежных ср-в, руб': krmo}
    df = pd.DataFrame(data)
    return df


df1 = get_datas()
df2 = get_ARAS()
df = df1.join(df2)
df = df.groupby('Покупатель').agg({'Поступление, рубли': 'sum', 'Курсовые разницы между отгрузкой и поступлением денежных ср-в, руб': 'sum'}).reset_index()
df = df.rename(columns={'Поступление, рубли': 'Поступление (руб)', 'Курсовые разницы между отгрузкой и поступлением денежных ср-в, руб': 'Прибыль за счёт изменения курса (руб)'})
df.to_csv("customers.csv")
